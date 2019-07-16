from openpyxl import load_workbook


wb = load_workbook('Financial Sample.xlsx')


print(wb.sheetnames)

ws1 = wb['Finances 2017']

print(ws1['B9'].value)

profit_total = 0
for column in list('L'):
    for row in range(2, 101):
        cell = column + str(row)
        profit_total += float(ws1[cell].value)

print(profit_total)


for row in range(2, ws1.max_row):
    cell = 'B' + str(row)
    print(ws1[cell].value)